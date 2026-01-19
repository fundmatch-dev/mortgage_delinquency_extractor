"""Excel formatting utilities for mortgage data output."""

import pandas as pd
from pathlib import Path
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelFormatter:
    """Formats and writes extracted mortgage data to Excel."""

    COLUMNS = [
        "filename",
        "series_name",
        "report_date",
        "beginning_balance",
        "ending_balance",
        "delinquency_category",
        "loan_count",
        "balance",
    ]

    def __init__(self):
        self.data_rows: list[dict] = []

    def add_document(self, doc_data: dict):
        """
        Add a document's data, flattening delinquency categories to long format.

        Args:
            doc_data: Dictionary with document metadata and delinquency list
        """
        metadata = {
            "filename": doc_data.get("filename"),
            "series_name": doc_data.get("series_name"),
            "report_date": doc_data.get("report_date"),
            "beginning_balance": doc_data.get("beginning_balance"),
            "ending_balance": doc_data.get("ending_balance"),
        }

        delinquency = doc_data.get("delinquency", [])

        if not delinquency:
            # Add a row with empty delinquency data
            row = {**metadata, "delinquency_category": None, "loan_count": None, "balance": None}
            self.data_rows.append(row)
        else:
            for cat in delinquency:
                row = {
                    **metadata,
                    "delinquency_category": cat.get("category"),
                    "loan_count": cat.get("count"),
                    "balance": cat.get("balance"),
                }
                self.data_rows.append(row)

    def add_document_from_model(self, doc):
        """Add document data from a validated Pydantic model."""
        doc_dict = doc.model_dump()
        doc_dict["delinquency"] = [
            {"category": d.category, "count": d.count, "balance": d.balance}
            for d in doc.delinquency
        ]
        self.add_document(doc_dict)

    def to_dataframe(self) -> pd.DataFrame:
        """Convert collected data to a pandas DataFrame."""
        df = pd.DataFrame(self.data_rows, columns=self.COLUMNS)
        return df

    def save_to_excel(self, output_path: str | Path, sheet_name: str = "Mortgage Data"):
        """
        Save data to Excel with formatting.

        Args:
            output_path: Path for the output Excel file
            sheet_name: Name of the worksheet
        """
        df = self.to_dataframe()
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        currency_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        number_format = '#,##0'

        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border

                if r_idx == 1:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                else:
                    # Format currency columns
                    col_name = self.COLUMNS[c_idx - 1]
                    if col_name in ["beginning_balance", "ending_balance", "balance"]:
                        cell.number_format = currency_format
                    elif col_name == "loan_count":
                        cell.number_format = number_format

        # Adjust column widths
        column_widths = {
            "filename": 15,
            "series_name": 18,
            "report_date": 12,
            "beginning_balance": 18,
            "ending_balance": 16,
            "delinquency_category": 22,
            "loan_count": 12,
            "balance": 16,
        }

        for idx, col_name in enumerate(self.COLUMNS, 1):
            ws.column_dimensions[chr(64 + idx)].width = column_widths.get(col_name, 15)

        # Freeze header row
        ws.freeze_panes = "A2"

        wb.save(output_path)
        return output_path

    def clear(self):
        """Clear all collected data."""
        self.data_rows = []

    def __len__(self):
        return len(self.data_rows)
